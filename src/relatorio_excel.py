from pathlib import Path
from datetime import date
import pandas as pd

from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, LineChart, Reference


def nome_relatorio_excel_com_data() -> str:
    return f"relatorio_{date.today().isoformat()}.xlsx"


def gerar_relatorio_excel(metricas: dict, caminho_saida: str) -> str:
    saida = Path(caminho_saida)
    saida.parent.mkdir(parents=True, exist_ok=True)

    df_det = metricas["df"]

    # Resumo
    ini = metricas["periodo_inicio"]
    fim = metricas["periodo_fim"]

    df_resumo = pd.DataFrame(
        {
            "Métrica": [
                "Data de geração",
                "Período (início)",
                "Período (fim)",
                "Número de vendas",
                "Quantidade total",
                "Faturamento total",
                "Ticket médio",
                "Produto campeão",
                "Dia recorde",
                "Faturamento dia recorde",
            ],
            "Valor": [
                date.today().isoformat(),
                ini.date().isoformat() if pd.notna(ini) else "",
                fim.date().isoformat() if pd.notna(fim) else "",
                metricas["numero_vendas"],
                metricas["quantidade_total"],
                metricas["faturamento_total"],
                metricas["ticket_medio"],
                metricas["produto_campeao"],
                str(metricas["dia_recorde"]) if metricas["dia_recorde"] else "",
                metricas["faturamento_dia_recorde"],
            ],
        }
    )

    # Por produto
    df_prod = metricas["faturamento_por_produto"].reset_index()
    df_prod.columns = ["produto", "faturamento"]
    df_prod["participacao_%"] = metricas["participacao_por_produto"].values

    # Por dia
    df_dia = metricas["faturamento_por_dia"].reset_index()
    df_dia.columns = ["data", "faturamento"]

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_prod.to_excel(writer, sheet_name="Por Produto", index=False)
        df_dia.to_excel(writer, sheet_name="Por Dia", index=False)
        df_det.to_excel(writer, sheet_name="Detalhado", index=False)

        wb = writer.book
        ws_resumo = wb["Resumo"]
        ws_prod = wb["Por Produto"]
        ws_dia = wb["Por Dia"]
        ws_det = wb["Detalhado"]

        def estilizar_header(ws):
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")

        for ws in (ws_resumo, ws_prod, ws_dia, ws_det):
            estilizar_header(ws)

        # Larguras
        ws_resumo.column_dimensions["A"].width = 26
        ws_resumo.column_dimensions["B"].width = 30

        ws_prod.column_dimensions["A"].width = 22
        ws_prod.column_dimensions["B"].width = 18
        ws_prod.column_dimensions["C"].width = 16

        ws_dia.column_dimensions["A"].width = 14
        ws_dia.column_dimensions["B"].width = 18

        # Formatos Resumo
        # Valores em Resumo estão na coluna B; vamos formatar por linha fixa (como escrevemos acima).
        # (Linha 1 é cabeçalho)
        # faturamento_total = linha 7 (B7), ticket médio = linha 8 (B8), faturamento dia recorde = linha 11 (B11)
        ws_resumo["B7"].number_format = 'R$ #,##0.00'
        ws_resumo["B8"].number_format = 'R$ #,##0.00'
        ws_resumo["B11"].number_format = 'R$ #,##0.00'

        # Formatos Por Produto
        for r in range(2, ws_prod.max_row + 1):
            ws_prod[f"B{r}"].number_format = 'R$ #,##0.00'
            ws_prod[f"C{r}"].number_format = '0.00"%"'

        # Formatos Por Dia
        for r in range(2, ws_dia.max_row + 1):
            ws_dia[f"B{r}"].number_format = 'R$ #,##0.00'

        # Formatos Detalhado (tentativa por nome de coluna)
        # Dataframe escrito com colunas: data, produto, quantidade, preco_unitario, total (se nessa ordem)
        # Mesmo se mudar, isso não quebra o Excel; só pode não formatar exatamente.
        # Aqui vamos formatar as colunas D e E como moeda (padrão do seu df).
        for r in range(2, ws_det.max_row + 1):
            ws_det[f"D{r}"].number_format = 'R$ #,##0.00'
            ws_det[f"E{r}"].number_format = 'R$ #,##0.00'

        # Gráfico Top 10 produtos (na aba Por Produto)
        max_top = min(10, ws_prod.max_row - 1)
        if max_top >= 1:
            data_ref = Reference(ws_prod, min_col=2, min_row=1, max_row=1 + max_top)
            cats_ref = Reference(ws_prod, min_col=1, min_row=2, max_row=1 + max_top)

            bar = BarChart()
            bar.title = "Top 10 Produtos por Faturamento"
            bar.y_axis.title = "Faturamento (R$)"
            bar.add_data(data_ref, titles_from_data=True)
            bar.set_categories(cats_ref)
            bar.height = 10
            bar.width = 20

            ws_prod.add_chart(bar, "E2")

        # Gráfico Faturamento por dia (na aba Por Dia)
        if ws_dia.max_row >= 3:
            data_ref = Reference(ws_dia, min_col=2, min_row=1, max_row=ws_dia.max_row)
            cats_ref = Reference(ws_dia, min_col=1, min_row=2, max_row=ws_dia.max_row)

            line = LineChart()
            line.title = "Faturamento por Dia"
            line.y_axis.title = "Faturamento (R$)"
            line.add_data(data_ref, titles_from_data=True)
            line.set_categories(cats_ref)
            line.height = 10
            line.width = 20

            ws_dia.add_chart(line, "D2")

    return str(saida)
